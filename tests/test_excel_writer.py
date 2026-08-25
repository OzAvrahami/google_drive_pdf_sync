"""Characterization tests for the current Excel export contract."""

from pathlib import Path

import os

import pytest
from openpyxl import Workbook, load_workbook

from app.models.document import Document
from app.writers import excel_writer
from app.writers.excel_writer import (
    _ALL_COLS,
    _DUP_COLOR,
    WorkbookLoadError,
    WorkbookStructureError,
    WorkbookWriteError,
    export_documents,
)


def _document(drive_id: str, **overrides) -> Document:
    values = {
        "drive_file_id": drive_id,
        "file_name": f"{drive_id}.pdf",
        "folder_path": "invoices/2026",
        "status": "approved",
        "supplier_name": "Example Supplier",
        "invoice_number": "INV-100",
        "invoice_date": "01/02/2026",
        "total": 250.0,
        "extracted_data": {"document_type": "tax_invoice"},
    }
    values.update(overrides)
    return Document(**values)


def _worksheet(path: Path):
    return load_workbook(path).active


def test_export_writes_current_column_structure_and_corrected_values(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    doc = _document(
        "drive-1",
        corrected_data={"supplier_name": "Corrected Supplier", "total": 275.0},
    )

    result = export_documents([doc], str(output))

    assert result.rows_written == 1
    assert result.written_ids == ("drive-1",)
    assert result.already_present_ids == ()

    ws = _worksheet(output)
    headers = [cell.value for cell in ws[1]]
    assert headers == _ALL_COLS
    row = [cell.value for cell in ws[2]]
    assert row[3] == "Corrected Supplier"
    assert row[6] == 275
    assert row[7] == "drive-1"


def test_export_skips_an_existing_drive_id(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    doc = _document("drive-1")

    assert export_documents([doc], str(output)).rows_written == 1
    duplicate = export_documents([doc], str(output))
    assert duplicate.rows_written == 0
    assert duplicate.already_present_ids == ("drive-1",)

    ws = _worksheet(output)
    assert ws.max_row == 2


def test_export_appends_new_drive_ids_to_existing_workbook(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"

    export_documents([_document("drive-1")], str(output))
    assert export_documents(
        [_document("drive-2", invoice_number="INV-200")], str(output)
    ).rows_written == 1

    ws = _worksheet(output)
    assert ws.max_row == 3
    assert [ws.cell(row=row, column=8).value for row in (2, 3)] == ["drive-1", "drive-2"]


def test_export_highlights_rows_with_duplicate_business_keys(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    first = _document("drive-1")
    second = _document("drive-2")

    assert export_documents([first, second], str(output)).rows_written == 2

    ws = _worksheet(output)
    for row in (2, 3):
        assert all(
            (cell.fill.fgColor.rgb or "").endswith(_DUP_COLOR)
            for cell in ws[row]
        )


def test_formula_leading_text_is_neutralized_but_numbers_remain_numeric(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    doc = _document(
        "drive-1",
        file_name="=SUM(A1:A2).pdf",
        folder_path="+cmd",
        supplier_name="@supplier",
        invoice_number="-1+2",
        invoice_date="01/02/2026",
        total=275.5,
        description="=unused",
    )

    export_documents([doc], str(output))

    row = [cell.value for cell in _worksheet(output)[2]]
    assert row[0] == "'=SUM(A1:A2).pdf"
    assert row[1] == "'+cmd"
    assert row[3] == "'@supplier"
    assert row[5] == "'-1+2"
    assert row[4] == "01/02/2026"
    assert row[6] == pytest.approx(275.5)
    assert isinstance(row[6], (int, float))


def test_ordinary_hebrew_and_supplier_text_is_unchanged(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    export_documents(
        [_document("drive-1", supplier_name="ספק רגיל", invoice_number="INV-10")],
        str(output),
    )
    row = [cell.value for cell in _worksheet(output)[2]]
    assert row[3] == "ספק רגיל"
    assert row[5] == "INV-10"


def test_corrupt_existing_workbook_fails_closed_and_preserves_original(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    original = b"not an xlsx workbook"
    output.write_bytes(original)

    with pytest.raises(WorkbookLoadError):
        export_documents([_document("drive-1")], str(output))

    assert output.read_bytes() == original


def test_readable_but_incompatible_workbook_fails_closed(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "חשבוניות"
    worksheet.append(["wrong", "headers"])
    workbook.save(output)
    original = output.read_bytes()

    with pytest.raises(WorkbookStructureError):
        export_documents([_document("drive-1")], str(output))

    assert output.read_bytes() == original


def test_missing_required_sheet_fails_closed(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.active.title = "Other"
    workbook.save(output)

    with pytest.raises(WorkbookStructureError):
        export_documents([_document("drive-1")], str(output))


def test_atomic_replacement_failure_preserves_original_and_cleans_temp(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output = tmp_path / "output.xlsx"
    export_documents([_document("drive-1")], str(output))
    original = output.read_bytes()

    def fail_replace(source, destination):
        raise PermissionError("workbook is open")

    monkeypatch.setattr(excel_writer.os, "replace", fail_replace)
    with pytest.raises(WorkbookWriteError):
        export_documents([_document("drive-2")], str(output))

    assert output.read_bytes() == original
    assert list(tmp_path.glob(".*.tmp.xlsx")) == []


def test_successful_write_uses_same_directory_temporary_replacement(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output = tmp_path / "nested" / "output.xlsx"
    calls: list[tuple[Path, Path]] = []
    real_replace = os.replace

    def observe_replace(source, destination):
        calls.append((Path(source), Path(destination)))
        return real_replace(source, destination)

    monkeypatch.setattr(excel_writer.os, "replace", observe_replace)
    export_documents([_document("drive-1")], str(output))

    assert calls
    temporary, destination = calls[0]
    assert temporary.parent == output.parent
    assert destination == output
    assert not temporary.exists()


def test_structured_result_deduplicates_requested_ids(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    first = _document("drive-1")
    duplicate_object = _document("drive-1")

    result = export_documents([first, duplicate_object], str(output))

    assert result.requested_ids == ("drive-1",)
    assert result.attempted_ids == ("drive-1",)
    assert result.written_ids == ("drive-1",)
    assert result.confirmed_ids == ("drive-1",)


def test_formula_safe_drive_id_remains_deduplicatable(tmp_path: Path) -> None:
    output = tmp_path / "output.xlsx"
    risky_id = "-synthetic-id"
    first = export_documents([_document(risky_id)], str(output))
    second = export_documents([_document(risky_id)], str(output))

    assert first.written_ids == (risky_id,)
    assert second.written_ids == ()
    assert second.already_present_ids == (risky_id,)
    assert _worksheet(output).cell(row=2, column=8).value == "'-synthetic-id"
