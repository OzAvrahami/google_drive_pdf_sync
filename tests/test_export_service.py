from __future__ import annotations

from pathlib import Path

from app.application.export_service import ExportOutcome, ExportService
from app.models.document import Document
from app.writers.excel_writer import ExportResult
from openpyxl import load_workbook


def document(document_id: str, status: str = "approved") -> Document:
    return Document(
        drive_file_id=document_id,
        file_name=f"{document_id}.pdf",
        folder_path="synthetic",
        status=status,
        supplier_name="Supplier",
        invoice_number=f"INV-{document_id}",
        invoice_date="25/08/2026",
        total=100.0,
    )


class Repository:
    def __init__(self, *documents: Document, fail_write: bool = False) -> None:
        self.documents = {item.drive_file_id: item for item in documents}
        self.fail_write = fail_write
        self.upsert_many_calls: list[list[Document]] = []
        self.events: list[str] = []

    def get_by_drive_id(self, document_id: str) -> Document | None:
        self.events.append(f"get:{document_id}")
        return self.documents.get(document_id)

    def upsert(self, item: Document) -> None:
        raise AssertionError("ExportService must use the bulk boundary")

    def upsert_many(self, items: list[Document]) -> None:
        self.events.append("upsert_many")
        self.upsert_many_calls.append(items)
        if self.fail_write:
            raise OSError("store unavailable")
        for item in items:
            self.documents[item.drive_file_id] = item


class Writer:
    def __init__(self, *, written=(), present=()) -> None:
        self.written = tuple(written)
        self.present = tuple(present)
        self.calls: list[tuple[list[Document], str]] = []

    def __call__(self, documents: list[Document], output_path: str) -> ExportResult:
        self.calls.append((documents, output_path))
        requested = tuple(item.drive_file_id for item in documents)
        return ExportResult(
            requested_ids=requested,
            attempted_ids=self.written,
            written_ids=self.written,
            already_present_ids=self.present,
            workbook_path=output_path,
            rows_written=len(self.written),
        )


def test_exports_only_reloaded_approved_selected_documents(tmp_path: Path) -> None:
    repository = Repository(
        document("approved"), document("processed", "processed")
    )
    writer = Writer(written=("approved",))
    service = ExportService(repository, tmp_path / "out.xlsx", writer=writer)

    result = service.export_selected(["approved", "processed", "missing"])

    assert result.eligible_ids == ("approved",)
    assert result.ineligible_ids == ("processed",)
    assert result.missing_ids == ("missing",)
    assert [item.drive_file_id for item in writer.calls[0][0]] == ["approved"]
    assert result.transitioned_ids == ("approved",)
    assert repository.documents["approved"].status == "exported"
    assert repository.documents["processed"].status == "processed"
    assert result.outcome is ExportOutcome.PARTIAL


def test_written_and_already_present_ids_are_the_only_transitions(tmp_path: Path) -> None:
    repository = Repository(document("new-row"), document("existing-row"), document("unconfirmed"))
    writer = Writer(written=("new-row",), present=("existing-row",))

    result = ExportService(
        repository, tmp_path / "out.xlsx", writer=writer
    ).export_selected(["new-row", "existing-row", "unconfirmed"])

    assert result.transitioned_ids == ("new-row", "existing-row")
    assert repository.documents["unconfirmed"].status == "approved"
    assert result.outcome is ExportOutcome.PARTIAL


def test_all_confirmed_ids_produce_success_and_one_bulk_write(tmp_path: Path) -> None:
    repository = Repository(document("one"), document("two"))
    writer = Writer(written=("one",), present=("two",))

    result = ExportService(
        repository, tmp_path / "out.xlsx", writer=writer
    ).export_selected(["one", "two"])

    assert result.outcome is ExportOutcome.SUCCEEDED
    assert result.exported_count == 2
    assert len(repository.upsert_many_calls) == 1
    assert [item.status for item in repository.upsert_many_calls[0]] == [
        "exported",
        "exported",
    ]


def test_status_store_failure_reports_partial_after_workbook_success(tmp_path: Path) -> None:
    repository = Repository(document("one"), fail_write=True)
    writer = Writer(written=("one",))

    result = ExportService(
        repository, tmp_path / "out.xlsx", writer=writer
    ).export_selected(["one"])

    assert result.written_ids == ("one",)
    assert result.transitioned_ids == ()
    assert result.outcome is ExportOutcome.PARTIAL
    assert "OSError" in result.status_persistence_error
    assert repository.documents["one"].status == "approved"


def test_no_eligible_document_never_calls_writer_or_store(tmp_path: Path) -> None:
    repository = Repository(document("processed", "processed"))
    writer = Writer()

    result = ExportService(
        repository, tmp_path / "out.xlsx", writer=writer
    ).export_selected(["processed", "missing"])

    assert result.outcome is ExportOutcome.NOTHING_TO_EXPORT
    assert result.ineligible_ids == ("processed",)
    assert result.missing_ids == ("missing",)
    assert writer.calls == []
    assert repository.upsert_many_calls == []


def test_service_reloads_every_stable_id_before_writer_call(tmp_path: Path) -> None:
    repository = Repository(document("one"), document("two"))
    writer = Writer(written=("one", "two"))

    ExportService(repository, tmp_path / "out.xlsx", writer=writer).export_selected(
        ["one", "two"]
    )

    assert repository.events == ["get:one", "get:two", "upsert_many"]


def test_real_workbook_success_is_reported_partial_when_status_store_fails(
    tmp_path: Path,
) -> None:
    repository = Repository(document("one"), fail_write=True)
    output = tmp_path / "out.xlsx"

    result = ExportService(repository, output).export_selected(["one"])

    assert result.outcome is ExportOutcome.PARTIAL
    assert result.written_ids == ("one",)
    assert result.transitioned_ids == ()
    assert repository.documents["one"].status == "approved"
    worksheet = load_workbook(output).active
    assert worksheet.cell(row=2, column=8).value == "one"
