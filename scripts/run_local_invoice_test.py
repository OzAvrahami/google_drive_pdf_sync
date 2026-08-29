"""
Local invoice pipeline test runner
===================================
Scans a local folder of downloaded PDFs, runs the invoice parser,
and writes results to an Excel file with Hebrew column headers.
Duplicate rows (same supplier + doc number + date + amount) are highlighted.

Usage (from the project root):
    python -m scripts.run_local_invoice_test

Edit LOCAL_PDF_ROOT and OUTPUT_EXCEL_PATH below before running.
"""

import sys
from pathlib import Path
from collections import Counter

# Allow imports from project root as 'app.*'
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.parsers.pdf_parser import extract_text_from_pdf
from app.parsers.pdf_layout import apply_positional_supplier_override
from app.parsers.invoice_parser import classify_document_type, parse_invoice_text

# ── Configuration ──────────────────────────────────────────────────────────────


#LOCAL_PDF_ROOT    = r"D:\code\google_drive_pdf_sync\downloads\ינואר 2026\חשבון עסקה בגין ינואר"
LOCAL_PDF_ROOT    = r"D:\code\google_drive_pdf_sync\downloads\ינואר 2026\חשבוניות מס"
#OUTPUT_EXCEL_PATH = r"D:\code\google_drive_pdf_sync\output\test\invoice_test_results.xlsx"
OUTPUT_EXCEL_PATH = r"D:\code\google_drive_pdf_sync\output\test\invoice_test_results_new.xlsx"

# ── PDF scanning ───────────────────────────────────────────────────────────────

def scan_pdfs(root: str) -> list[dict]:
    root_path = Path(root)
    records = []
    for pdf_path in sorted(root_path.rglob("*.pdf")):
        records.append({
            "file_name": pdf_path.name,
            "full_path": str(pdf_path),
        })
    return records


# ── Per-file processing ────────────────────────────────────────────────────────

_PROCESSABLE = {"חשבון עסקה", "חשבונית מס", "דרישת תשלום"}

_TYPE_LABELS = {
    "transaction_invoice": "חשבון עסקה",
    "tax_invoice":         "חשבונית מס",
    "payment_request":     "דרישת תשלום",
}

def process_pdf(meta: dict) -> dict:
    row = {
        "שם קובץ":    meta["file_name"],
        "סוג קובץ":   "",
        "שם ספק":     "",
        "תאריך מסמך": "",
        "מספר מסמך":  "",
        "סכום":       "",
        # internal only — stripped before Excel write
        "_status":    "unrecognized_document_type",
        "_full_path": meta["full_path"],
    }

    try:
        text = extract_text_from_pdf(meta["full_path"])
        doc_type = classify_document_type(text)

        if doc_type is None:
            return row

        if doc_type not in _PROCESSABLE:
            row["_status"] = f"skipped_{doc_type}"
            return row

        row["_status"] = "processed"
        parsed = parse_invoice_text(text)
        if parsed:
            apply_positional_supplier_override(
                parsed,
                text,
                pdf_path=meta["full_path"],
            )
            row["סוג קובץ"]   = _TYPE_LABELS.get(parsed.get("document_type", ""), doc_type)
            row["שם ספק"]     = parsed.get("business_name")  or ""
            row["תאריך מסמך"] = parsed.get("invoice_date")   or ""
            row["מספר מסמך"]  = parsed.get("invoice_number") or ""
            amount = parsed.get("amount")
            row["סכום"] = amount if amount is not None else ""

    except Exception as exc:
        row["_status"] = f"parse_failed: {type(exc).__name__}: {exc}"

    return row


# ── Duplicate detection ────────────────────────────────────────────────────────

def find_duplicate_rows(rows: list[dict]) -> set[int]:
    """
    Return the set of row indices (0-based) that share the same
    (שם ספק, מספר מסמך, תאריך מסמך, סכום) key.
    Rows where all four key fields are empty are excluded.
    """
    key_counts: Counter = Counter()
    keys = []
    for row in rows:
        k = (row["שם ספק"], row["מספר מסמך"], row["תאריך מסמך"], str(row["סכום"]))
        keys.append(k)
        if any(v for v in k):          # at least one field non-empty
            key_counts[k] += 1

    return {i for i, k in enumerate(keys) if key_counts[k] > 1 and any(v for v in k)}


# ── Excel export ───────────────────────────────────────────────────────────────

_EXCEL_COLS = ["שם קובץ", "סוג קובץ", "שם ספק", "תאריך מסמך", "מספר מסמך", "סכום"]
_DUP_FILL_COLOR = "FFC7CE"   # standard Excel "bad" light-red highlight


def write_excel(rows: list[dict], output_path: str, dup_indices: set[int]) -> None:
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandas is required: pip install pandas openpyxl")

    df = pd.DataFrame(rows, columns=_EXCEL_COLS)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False, sheet_name="תוצאות")

    if dup_indices:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        fill = PatternFill(start_color=_DUP_FILL_COLOR, end_color=_DUP_FILL_COLOR, fill_type="solid")
        wb = load_workbook(output_path)
        ws = wb.active
        for idx in dup_indices:
            excel_row = idx + 2          # +1 for header, +1 for 1-based index
            for cell in ws[excel_row]:
                cell.fill = fill
        wb.save(output_path)


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if not Path(LOCAL_PDF_ROOT).exists():
        print(f"ERROR: LOCAL_PDF_ROOT does not exist: {LOCAL_PDF_ROOT}")
        sys.exit(1)

    print(f"Scanning: {LOCAL_PDF_ROOT}")
    pdfs = scan_pdfs(LOCAL_PDF_ROOT)
    print(f"Found {len(pdfs)} PDF(s)\n")

    rows = []
    for meta in pdfs:
        print(f"  {meta['file_name']} ... ", end="", flush=True)
        row = process_pdf(meta)
        rows.append(row)
        print(row["_status"])

    dup_indices = find_duplicate_rows(rows)
    if dup_indices:
        print(f"\n  ⚠  {len(dup_indices)} duplicate row(s) detected — will be highlighted in red")

    # Strip internal fields before export
    export_rows = [{k: v for k, v in r.items() if not k.startswith("_")} for r in rows]
    write_excel(export_rows, OUTPUT_EXCEL_PATH, dup_indices)

    # Summary
    statuses = [r["_status"] for r in rows]
    processed = sum(1 for s in statuses if s == "processed")
    skipped   = sum(1 for s in statuses if s.startswith("skipped"))
    failed    = sum(1 for s in statuses if s.startswith("parse_failed"))
    other     = len(rows) - processed - skipped - failed

    print(f"\n── Summary ───────────────────────────")
    print(f"  Total PDFs : {len(rows)}")
    print(f"  Processed  : {processed}")
    print(f"  Skipped    : {skipped}")
    print(f"  Failed     : {failed}")
    if other:
        print(f"  Other      : {other}")
    print(f"\nExcel written to: {OUTPUT_EXCEL_PATH}")


if __name__ == "__main__":
    main()
