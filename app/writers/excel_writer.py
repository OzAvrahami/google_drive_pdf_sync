"""Fail-closed, formula-safe Excel output for Panda documents.

The workbook contract intentionally remains the legacy Panda contract: one
``חשבוניות`` sheet, seven accountant-facing columns, and the existing
``drive_file_id`` deduplication column. Existing compatible workbooks are
loaded before a complete replacement workbook is written beside the target and
atomically installed.
"""

from __future__ import annotations

from dataclasses import dataclass, field
import logging
import os
from pathlib import Path
import tempfile
from types import MappingProxyType
from typing import Mapping, Optional

logger = logging.getLogger(__name__)

_SHEET_NAME = "חשבוניות"
_DISPLAY_COLS = [
    "שם קובץ",
    "נתיב",
    "סוג מסמך",
    "שם ספק",
    "תאריך מסמך",
    "מספר מסמך",
    "סכום",
]
_ALL_COLS = _DISPLAY_COLS + ["drive_file_id"]
_DUP_COLOR = "FFC7CE"

_TYPE_LABELS = {
    "transaction_invoice": "חשבון עסקה",
    "tax_invoice": "חשבונית מס",
    "payment_request": "דרישת תשלום",
}


class ExcelExportError(RuntimeError):
    """Base class for safe workbook failures."""

    def __init__(self, path: str | Path, reason: str) -> None:
        self.path = Path(path)
        self.reason = reason
        super().__init__(f"Excel export could not safely use {self.path.name}: {reason}")


class WorkbookLoadError(ExcelExportError):
    """The existing workbook could not be read without risking overwrite."""


class WorkbookStructureError(ExcelExportError):
    """The existing workbook is readable but not a compatible Panda workbook."""


class WorkbookWriteError(ExcelExportError):
    """A replacement workbook could not be built or installed safely."""


@dataclass(frozen=True, slots=True)
class ExportResult:
    requested_ids: tuple[str, ...]
    attempted_ids: tuple[str, ...]
    written_ids: tuple[str, ...]
    already_present_ids: tuple[str, ...]
    failed_ids: tuple[str, ...] = ()
    failures: Mapping[str, str] = field(default_factory=dict)
    workbook_path: str = ""
    rows_written: int = 0

    def __post_init__(self) -> None:
        object.__setattr__(self, "failures", MappingProxyType(dict(self.failures)))

    @property
    def confirmed_ids(self) -> tuple[str, ...]:
        return self.written_ids + tuple(
            value for value in self.already_present_ids if value not in self.written_ids
        )

    @property
    def is_complete(self) -> bool:
        return not self.failed_ids and set(self.confirmed_ids) == set(self.requested_ids)


def export_documents(docs: list, output_path: str) -> ExportResult:
    """Export supplied documents and report exact per-Drive-ID outcomes.

    Eligibility and status changes belong to the application export service.
    This writer only performs workbook I/O and Drive-ID deduplication.
    """
    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError("pandas is required: pip install pandas openpyxl") from exc

    target = Path(output_path)
    requested = tuple(dict.fromkeys(str(doc.drive_file_id) for doc in docs))
    existing_df = _load_existing(target, _ALL_COLS)
    existing_ids = _existing_ids(existing_df, "drive_file_id")

    rows: list[dict] = []
    written: list[str] = []
    already_present: list[str] = []
    seen = set(existing_ids)
    for doc in docs:
        fid = str(doc.drive_file_id)
        if fid in seen:
            if fid in existing_ids and fid not in already_present:
                already_present.append(fid)
            continue
        seen.add(fid)
        written.append(fid)
        total = doc.effective("total")
        rows.append(
            {
                "שם קובץ": _safe_excel_text(doc.file_name),
                "נתיב": _safe_excel_text(doc.folder_path),
                "סוג מסמך": _safe_excel_text(
                    _type_label(doc.extracted_data.get("document_type"))
                ),
                "שם ספק": _safe_excel_text(doc.effective("supplier_name") or ""),
                "תאריך מסמך": _safe_excel_text(doc.effective("invoice_date") or ""),
                "מספר מסמך": _safe_excel_text(doc.effective("invoice_number") or ""),
                "סכום": total if total is not None else "",
                "drive_file_id": _safe_excel_text(fid),
            }
        )

    if rows:
        new_df = pd.DataFrame(rows, columns=_ALL_COLS)
        combined = (
            pd.concat([existing_df, new_df], ignore_index=True)
            if not existing_df.empty
            else new_df
        )
        _write_dataframe_atomic(combined, target, _ALL_COLS)

    result = ExportResult(
        requested_ids=requested,
        attempted_ids=requested,
        written_ids=tuple(written),
        already_present_ids=tuple(already_present),
        workbook_path=str(target),
        rows_written=len(written),
    )
    logger.info(
        "Excel export confirmed %d new and %d existing row(s) in %s",
        len(written),
        len(already_present),
        target,
    )
    return result


def append_records(records: list, output_path: str) -> None:
    """Append legacy ``InvoiceRecord`` objects with the same safety boundary."""
    try:
        import pandas as pd
    except ImportError as exc:
        raise ImportError("pandas is required: pip install pandas openpyxl") from exc

    legacy_display = [
        "שם קובץ",
        "סוג קובץ",
        "שם ספק",
        "תאריך מסמך",
        "מספר מסמך",
        "סכום",
    ]
    legacy_all = legacy_display + ["file_id"]
    target = Path(output_path)
    existing = _load_existing(target, legacy_all)
    existing_ids = _existing_ids(existing, "file_id")
    rows: list[dict] = []
    seen = set(existing_ids)
    for record in records:
        fid = str(getattr(record, "file_id", "") or "")
        if fid and fid in seen:
            continue
        if fid:
            seen.add(fid)
        amount = getattr(record, "amount", None)
        rows.append(
            {
                "שם קובץ": _safe_excel_text(getattr(record, "file_name", "") or ""),
                "סוג קובץ": _safe_excel_text(
                    _type_label(getattr(record, "document_type", None))
                ),
                "שם ספק": _safe_excel_text(getattr(record, "business_name", "") or ""),
                "תאריך מסמך": _safe_excel_text(
                    getattr(record, "invoice_date", "") or ""
                ),
                "מספר מסמך": _safe_excel_text(
                    getattr(record, "invoice_number", "") or ""
                ),
                "סכום": amount if amount is not None else "",
                "file_id": _safe_excel_text(fid),
            }
        )
    if not rows:
        return
    new_df = pd.DataFrame(rows, columns=legacy_all)
    combined = (
        pd.concat([existing, new_df], ignore_index=True)
        if not existing.empty
        else new_df
    )
    _write_dataframe_atomic(combined, target, legacy_all)


def _safe_excel_text(value: object) -> object:
    """Neutralize formula-leading text while preserving non-string cell types."""
    if not isinstance(value, str) or not value:
        return value
    candidate = value.lstrip(" \t\r\n")
    if candidate.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _type_label(doc_type: Optional[str]) -> str:
    if not doc_type:
        return ""
    return _TYPE_LABELS.get(doc_type, doc_type)


def _load_existing(path: str | Path, expected_columns: list[str]):
    import pandas as pd
    from openpyxl import load_workbook

    path = Path(path)
    if not path.exists():
        return pd.DataFrame(columns=expected_columns)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise WorkbookLoadError(
            path, f"the existing workbook is unreadable ({type(exc).__name__})"
        ) from exc
    try:
        if _SHEET_NAME not in workbook.sheetnames:
            raise WorkbookStructureError(path, f"required sheet '{_SHEET_NAME}' is missing")
        worksheet = workbook[_SHEET_NAME]
        headers = [
            cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
    except WorkbookStructureError:
        raise
    except Exception as exc:
        raise WorkbookStructureError(
            path, "the header row could not be inspected"
        ) from exc
    finally:
        workbook.close()
    if headers != expected_columns:
        raise WorkbookStructureError(
            path, "the header columns do not match the Panda workbook contract"
        )
    try:
        frame = pd.read_excel(path, sheet_name=_SHEET_NAME, dtype=str)
    except Exception as exc:
        raise WorkbookLoadError(
            path, f"the existing sheet is unreadable ({type(exc).__name__})"
        ) from exc
    return frame[expected_columns].fillna("")


def _existing_ids(frame, column: str) -> set[str]:
    if frame.empty:
        return set()
    result: set[str] = set()
    for value in frame[column].tolist():
        text = str(value)
        if text.startswith("'") and text[1:].lstrip(" \t\r\n").startswith(
            ("=", "+", "-", "@")
        ):
            text = text[1:]
        if text.strip():
            result.add(text)
    return result


def _write_dataframe_atomic(frame, target: Path, columns: list[str]) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    duplicate_indices = _duplicate_indices(frame, columns)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{target.stem}.", suffix=".tmp.xlsx", dir=target.parent
    )
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        frame.to_excel(temporary, index=False, sheet_name=_SHEET_NAME)
        _format_workbook(temporary, duplicate_indices)
        _validate_written_workbook(temporary, columns)
        os.replace(temporary, target)
    except ExcelExportError:
        raise
    except Exception as exc:
        raise WorkbookWriteError(
            target, f"the replacement could not be completed ({type(exc).__name__})"
        ) from exc
    finally:
        try:
            temporary.unlink(missing_ok=True)
        except OSError:
            logger.warning("Could not remove temporary Excel file %s", temporary)


def _duplicate_indices(frame, columns: list[str]) -> set[int]:
    required = {"שם ספק", "מספר מסמך", "תאריך מסמך", "סכום"}
    if not required.issubset(columns):
        return set()
    keys = ["שם ספק", "מספר מסמך", "תאריך מסמך", "סכום"]
    has_content = frame[keys].apply(
        lambda row: any(str(value).strip() for value in row), axis=1
    )
    tuples = frame[keys].apply(tuple, axis=1)
    frequencies = tuples.value_counts()
    mask = has_content & tuples.map(lambda key: frequencies[key] > 1)
    return set(frame.index[mask].tolist())


def _validate_written_workbook(path: Path, expected_columns: list[str]) -> None:
    from openpyxl import load_workbook

    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        if _SHEET_NAME not in workbook.sheetnames:
            raise ValueError("sheet missing after write")
        worksheet = workbook[_SHEET_NAME]
        headers = [
            cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        ]
        if headers != expected_columns:
            raise ValueError("headers differ after write")
    except Exception as exc:
        raise WorkbookWriteError(path, "replacement validation failed") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _format_workbook(path: str | Path, duplicate_indices: set[int]) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    fill = PatternFill(start_color=_DUP_COLOR, end_color=_DUP_COLOR, fill_type="solid")
    workbook = load_workbook(path)
    worksheet = workbook[_SHEET_NAME]
    for index in duplicate_indices:
        for cell in worksheet[index + 2]:
            cell.fill = fill
    for column_index, cells in enumerate(worksheet.columns, start=1):
        max_length = max((len(str(cell.value or "")) for cell in cells), default=10)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max_length + 4, 55
        )
    workbook.save(path)
    workbook.close()
